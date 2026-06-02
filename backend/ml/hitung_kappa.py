# ================================================================
# HITUNG COHEN'S KAPPA — ASPIRALYTICA IAA
# ================================================================
# Cara pakai:
#   1. Pastikan file Excel sudah diisi oleh KEDUA anotator
#   2. Simpan file Excel dengan nama yang sama
#   3. Jalankan: python hitung_kappa.py
#   4. Hasil kappa akan muncul di terminal DAN disimpan ke
#      file: HASIL_KAPPA_IAA.txt
# ================================================================

import openpyxl
import sys
import os
from collections import Counter
from datetime import datetime

try:
    from sklearn.metrics import cohen_kappa_score, confusion_matrix, classification_report
    SKLEARN_OK = True
except ImportError:
    SKLEARN_OK = False
    print("⚠️  sklearn tidak ditemukan. Install dengan: pip install scikit-learn")

# ── KONFIGURASI ──────────────────────────────────────────────────
# Ganti nama file jika berbeda
EXCEL_FILE = "LEMBAR_ANOTASIII.xlsx"

SHEET_A = "🅰️ ANOTATOR A"
SHEET_B = "🅱️ ANOTATOR B"

# Posisi data di sheet (sesuai file yang dibuat)
SENT_START_ROW   = 3     # baris pertama data sentimen
SENT_TOTAL       = 81    # jumlah teks sentimen
SENT_COL_A       = 3     # kolom C = Label Sentimen Anotator A
SENT_COL_B       = 3     # kolom C = Label Sentimen Anotator B

# Intent: mulai setelah sentimen + 2 baris header
INTENT_START_ROW = SENT_START_ROW + SENT_TOTAL + 2 + 1  # row 86 + 1 header = 87
INTENT_TOTAL     = 90    # jumlah teks intent
INTENT_COL_A     = 4     # kolom D = Label Intent Anotator A
INTENT_COL_B     = 4     # kolom D = Label Intent Anotator B

# ── VALIDASI LABEL ───────────────────────────────────────────────
VALID_SENT   = {"positif", "negatif", "netral"}
VALID_INTENT = {"keluhan", "permintaan", "saran", "apresiasi", "darurat"}

# ── WARNA TERMINAL ───────────────────────────────────────────────
def green(s):  return f"\033[92m{s}\033[0m"
def red(s):    return f"\033[91m{s}\033[0m"
def yellow(s): return f"\033[93m{s}\033[0m"
def bold(s):   return f"\033[1m{s}\033[0m"
def cyan(s):   return f"\033[96m{s}\033[0m"

# ── INTERPRETASI KAPPA ───────────────────────────────────────────
def interpret_kappa(k):
    if k < 0.00: return "Poor", "❌ TIDAK DAPAT DITERIMA", red
    if k < 0.20: return "Slight", "❌ TIDAK DAPAT DITERIMA", red
    if k < 0.40: return "Fair", "⚠️  LEMAH — perlu perbaikan guideline", yellow
    if k < 0.61: return "Moderate", "⚠️  MINIMUM — dapat diterima dengan catatan", yellow
    if k < 0.80: return "Substantial", "✅ BAIK — cukup untuk SINTA 4", green
    return "Almost Perfect", "✅ SANGAT BAIK — standar internasional", green

# ── BACA EXCEL ───────────────────────────────────────────────────
def read_labels(ws, start_row, total_rows, col_a_or_b):
    labels = []
    empty_rows = []
    invalid_rows = []
    for i in range(total_rows):
        row = start_row + i
        val = ws.cell(row=row, column=col_a_or_b).value
        if val is None or str(val).strip() == "" or str(val).strip() == "—":
            empty_rows.append(row)
            labels.append(None)
        else:
            val_clean = str(val).strip().lower()
            labels.append(val_clean)
    return labels, empty_rows

def check_file():
    if not os.path.exists(EXCEL_FILE):
        # Coba cari di direktori yang sama dengan script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        alt_path = os.path.join(script_dir, EXCEL_FILE)
        if os.path.exists(alt_path):
            return alt_path
        print(red(f"\n❌  File tidak ditemukan: {EXCEL_FILE}"))
        print(f"   Pastikan file Excel ada di folder yang sama dengan script ini.")
        print(f"   Folder script: {script_dir}")
        sys.exit(1)
    return EXCEL_FILE

# ── HITUNG KAPPA MANUAL (tanpa sklearn) ──────────────────────────
def kappa_manual(labels_a, labels_b, classes):
    n = len(labels_a)
    if n == 0: return 0.0
    
    agree = sum(1 for a, b in zip(labels_a, labels_b) if a == b)
    po = agree / n  # observed agreement
    
    count_a = Counter(labels_a)
    count_b = Counter(labels_b)
    
    pe = sum((count_a.get(c, 0) / n) * (count_b.get(c, 0) / n) for c in classes)
    
    if pe == 1.0: return 1.0
    return (po - pe) / (1 - pe)

# ── MAIN ─────────────────────────────────────────────────────────
def main():
    print("\n" + "="*65)
    print(bold(cyan("  HITUNG COHEN'S KAPPA — ASPIRALYTICA IAA")))
    print("="*65)
    print(f"  File    : {EXCEL_FILE}")
    print(f"  Waktu   : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*65 + "\n")

    file_path = check_file()

    print(f"📂  Membaca file: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(red(f"❌  Gagal membuka file: {e}"))
        sys.exit(1)

    if SHEET_A not in wb.sheetnames or SHEET_B not in wb.sheetnames:
        print(red(f"❌  Sheet tidak ditemukan!"))
        print(f"   Sheet yang ada: {wb.sheetnames}")
        sys.exit(1)

    ws_a = wb[SHEET_A]
    ws_b = wb[SHEET_B]
    print(green(f"✅  File berhasil dibaca. Sheet A dan B ditemukan.\n"))

    results = []  # untuk disimpan ke file txt

    # ════════════════════════════════════════════════════════════════
    # SENTIMEN
    # ════════════════════════════════════════════════════════════════
    print(bold("─── SENTIMEN ──────────────────────────────────────────────────"))

    sent_a, empty_a = read_labels(ws_a, SENT_START_ROW, SENT_TOTAL, SENT_COL_A)
    sent_b, empty_b = read_labels(ws_b, SENT_START_ROW, SENT_TOTAL, SENT_COL_B)

    # Cek kelengkapan
    if empty_a:
        print(yellow(f"⚠️  Anotator A: {len(empty_a)} baris sentimen masih kosong (baris: {empty_a[:5]}{'...' if len(empty_a)>5 else ''})"))
    if empty_b:
        print(yellow(f"⚠️  Anotator B: {len(empty_b)} baris sentimen masih kosong (baris: {empty_b[:5]}{'...' if len(empty_b)>5 else ''})"))

    # Hanya hitung pada baris yang keduanya terisi
    sent_pairs = [(a, b) for a, b in zip(sent_a, sent_b) if a is not None and b is not None]
    
    if len(sent_pairs) == 0:
        print(red("❌  Tidak ada data sentimen yang bisa dihitung. Pastikan kedua anotator sudah mengisi label."))
    else:
        la_s = [p[0] for p in sent_pairs]
        lb_s = [p[1] for p in sent_pairs]
        
        n_sent = len(sent_pairs)
        agree_sent = sum(1 for a, b in sent_pairs if a == b)
        pct_agree_sent = agree_sent / n_sent * 100

        if SKLEARN_OK:
            kappa_sent = cohen_kappa_score(la_s, lb_s)
        else:
            kappa_sent = kappa_manual(la_s, lb_s, list(VALID_SENT))

        cat_s, msg_s, color_s = interpret_kappa(kappa_sent)

        print(f"\n  Total teks dievaluasi : {n_sent} / {SENT_TOTAL}")
        print(f"  Agreement (sama)      : {agree_sent} teks ({pct_agree_sent:.1f}%)")
        print(f"  Disagreement (beda)   : {n_sent - agree_sent} teks ({100-pct_agree_sent:.1f}%)")
        print(f"  Cohen's Kappa         : {bold(color_s(f'κ = {kappa_sent:.4f}'))}")
        print(f"  Interpretasi          : {color_s(cat_s + ' — ' + msg_s)}")

        # Distribusi per kelas
        print(f"\n  Distribusi label Anotator A: {dict(Counter(la_s))}")
        print(f"  Distribusi label Anotator B: {dict(Counter(lb_s))}")

        # Disagreement detail
        disagree_sent = [(i+1, a, b) for i,(a,b) in enumerate(sent_pairs) if a != b]
        if disagree_sent:
            print(f"\n  {yellow('Detail Disagreement Sentimen')} ({len(disagree_sent)} kasus):")
            for no, a, b in disagree_sent[:10]:
                print(f"    Teks #{no:3d} → A: {a:10s} | B: {b}")
            if len(disagree_sent) > 10:
                print(f"    ... dan {len(disagree_sent)-10} kasus lainnya")

        results.append(("SENTIMEN", n_sent, agree_sent, pct_agree_sent, kappa_sent, cat_s, disagree_sent))

    # ════════════════════════════════════════════════════════════════
    # INTENT
    # ════════════════════════════════════════════════════════════════
    print(f"\n{bold('─── INTENT ────────────────────────────────────────────────────')}")

    # Intent start row: setelah sentimen section + 2 baris header section
    intent_row_start = SENT_START_ROW + SENT_TOTAL + 2 + 1

    int_a, empty_ia = read_labels(ws_a, intent_row_start, INTENT_TOTAL, INTENT_COL_A)
    int_b, empty_ib = read_labels(ws_b, intent_row_start, INTENT_TOTAL, INTENT_COL_B)

    if empty_ia:
        print(yellow(f"⚠️  Anotator A: {len(empty_ia)} baris intent masih kosong"))
    if empty_ib:
        print(yellow(f"⚠️  Anotator B: {len(empty_ib)} baris intent masih kosong"))

    intent_pairs = [(a, b) for a, b in zip(int_a, int_b) if a is not None and b is not None]

    if len(intent_pairs) == 0:
        print(red("❌  Tidak ada data intent yang bisa dihitung."))
    else:
        la_i = [p[0] for p in intent_pairs]
        lb_i = [p[1] for p in intent_pairs]

        n_int = len(intent_pairs)
        agree_int = sum(1 for a, b in intent_pairs if a == b)
        pct_agree_int = agree_int / n_int * 100

        if SKLEARN_OK:
            kappa_int = cohen_kappa_score(la_i, lb_i)
        else:
            kappa_int = kappa_manual(la_i, lb_i, list(VALID_INTENT))

        cat_i, msg_i, color_i = interpret_kappa(kappa_int)

        print(f"\n  Total teks dievaluasi : {n_int} / {INTENT_TOTAL}")
        print(f"  Agreement (sama)      : {agree_int} teks ({pct_agree_int:.1f}%)")
        print(f"  Disagreement (beda)   : {n_int - agree_int} teks ({100-pct_agree_int:.1f}%)")
        print(f"  Cohen's Kappa         : {bold(color_i(f'κ = {kappa_int:.4f}'))}")
        print(f"  Interpretasi          : {color_i(cat_i + ' — ' + msg_i)}")

        print(f"\n  Distribusi label Anotator A: {dict(Counter(la_i))}")
        print(f"  Distribusi label Anotator B: {dict(Counter(lb_i))}")

        disagree_int = [(i+1, a, b) for i,(a,b) in enumerate(intent_pairs) if a != b]
        if disagree_int:
            print(f"\n  {yellow('Detail Disagreement Intent')} ({len(disagree_int)} kasus):")
            for no, a, b in disagree_int[:10]:
                print(f"    Teks #{no:3d} → A: {a:12s} | B: {b}")
            if len(disagree_int) > 10:
                print(f"    ... dan {len(disagree_int)-10} kasus lainnya")

        results.append(("INTENT", n_int, agree_int, pct_agree_int, kappa_int, cat_i, disagree_int))

    # ════════════════════════════════════════════════════════════════
    # RINGKASAN AKHIR
    # ════════════════════════════════════════════════════════════════
    print(f"\n{'='*65}")
    print(bold(cyan("  RINGKASAN AKHIR")))
    print(f"{'='*65}")

    output_lines = []
    output_lines.append("HASIL INTER-ANNOTATOR AGREEMENT (IAA) — ASPIRALYTICA")
    output_lines.append(f"Tanggal: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    output_lines.append("="*65)

    for task, n, agree, pct, kappa, cat, disagree in results:
        _, msg, color = interpret_kappa(kappa)
        line = f"\n  {task:10s}  |  n={n}  |  Agreement={pct:.1f}%  |  κ={kappa:.4f}  |  {cat}"
        print(color(line))
        output_lines.append(line)

    print(f"\n{'='*65}")

    # Template untuk paper
    if len(results) == 2:
        _, _, _, pct_s, kappa_s, cat_s, _ = results[0]
        _, _, _, pct_i, kappa_i, cat_i, _ = results[1]
        
        template = f"""
  📝  TEMPLATE UNTUK PAPER (Section 2.2):
  ─────────────────────────────────────────────────────────────
  "Annotation reliability was assessed through inter-annotator
  agreement (IAA) on a stratified 20% subset of the corpus
  (n = 81 sentiment samples, n = 90 intent samples),
  independently labeled by a second annotator following the
  same written guidelines. Cohen's Kappa coefficients of
  κ = {kappa_s:.2f} for sentiment and κ = {kappa_i:.2f} for intent
  indicate {cat_s.lower()} agreement (Landis and Koch, 1977),
  confirming the reliability of the annotation scheme.
  Disagreements were resolved through adjudication by the
  lead annotator."
  ─────────────────────────────────────────────────────────────"""
        print(bold(template))
        output_lines.append(template)

    # Simpan ke file txt
    output_file = "HASIL_KAPPA_IAA.txt"
    try:
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(output_lines))
        print(f"\n  💾  Hasil disimpan ke: {green(output_file)}")
    except Exception as e:
        print(yellow(f"\n  ⚠️  Gagal menyimpan file: {e}"))

    print(f"\n{'='*65}\n")

if __name__ == "__main__":
    main()